import openpyxl
import os
import datetime
from Data import NewSuperObject
from openpyxl.styles import Font

def ReturnList(path:str):
    workbook = openpyxl.load_workbook(path)   
    sheets = workbook.sheetnames
    workbook.close()
    return sheets

def UploadingFile(path:str, _sheet:str):

    
    if os.path.isfile(path):
        workbook = openpyxl.load_workbook(path)
        if _sheet == "Создать нойвый лист":
            sheet = workbook.create_sheet("Новый лист")
        else:
            sheet = workbook[_sheet]
            
    if os.path.isdir(path):
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Новый лист")

    row = sheet.max_row+1
    
    for sup in NewSuperObject:
        sheet.cell(row=row, column=1).value = sup.CadNumber
        sheet.cell(row=row, column=2).value = sup.type_of_object
        if len(sup.oks)>0:
            temp_zu = ", ".join(sup.oks)
            sheet.cell(row=row, column=3).value = temp_zu
        else:
            sheet.cell(row=row, column=3).value ="не найдено"
            
        sheet.cell(row=row, column=4).value = "ОКС на ЗУ"
        row+=1
        sheet.cell(row=row, column=1).value = "Площадь и погрешность КПТ"
        sheet.cell(row=row, column=2).value = sup.square_kpt
        sheet.cell(row=row, column=3).value = sup.error_determining_area
        sheet.cell(row=row, column=4).value = sup.presence_coordinates
        row+=1
        sheet.cell(row=row, column=1).value = "Площадь по координатам"
        sheet.cell(row=row, column=2).value = sup.square_mif
        row+=1
        sheet.cell(row=row, column=1).value = "% отклонения от КПТ"
        sheet.cell(row=row, column=2).value = sup.deviations
        
        if sup.deviations != 'данные отсутствуют':
            if sup.deviations < 0:
                sheet.cell(row=row,column=3).value= "уменьшение"
            else:
                sheet.cell(row=row,column=3).value= "увеличение"
        else:
            sheet.cell(row=row,column=3).value= sup.deviations
        
        row+=1
        
        if sup.x_kpt and sup.x_kpt != 'данные отсутствуют': # кпт координаты 
            for ixsupkpt, xsupkpt in enumerate(sup.x_kpt):
                sheet.cell(row=row, column=2).value = sup.point_number[ixsupkpt]
                sheet.cell(row=row, column=3).value = xsupkpt
                sheet.cell(row=row, column=4).value = sup.y_kpt[ixsupkpt]
                row+=1
        else:
            sheet.cell(row=row, column=2).value = "данные отсутствуют"
            sheet.cell(row=row, column=3).value = "данные отсутствуют"            
            sheet.cell(row=row, column=4).value = "данные отсутствуют"
            row+=1
            
        if sup.finaly_row: # финульные строки 
            for ifin, fin in enumerate(sup.finaly_row):
                if fin.number_row == None:
                    sheet.cell(row=row, column=5).value = ""
                else:
                    sheet.cell(row=row, column=5).value = fin.number_row
                
                sheet.cell(row=row, column=6).value = fin.x
                sheet.cell(row=row, column=7).value = fin.y
                sheet.cell(row=row, column=8).value = fin.error_rate
                sheet.cell(row=row, column=9).value = fin.method_determining_point
                sheet.cell(row=row, column=10).value = fin.source
                if fin.source != "ЕГРН":   
                    sheet.cell(row=row, column=11).value = "Новая"
                    for number in range(6,11):
                        sheet.cell(row=row, column=number).font = Font(bold=True)
                else:
                    sheet.cell(row=row, column=11).value = ""
                    
                if fin.adjacent_codastres:
                    column = 12
                    for adjacent in fin.adjacent_codastres:
                        sheet.cell(row=row, column=column).value = adjacent
                        column+=1
                row+=1
                
                
            if sup.finaly_row[0].number_row == None:
                sheet.cell(row=row, column=5).value = ""
            else:
                sheet.cell(row=row, column=5).value = sup.finaly_row[0].number_row
            
            sheet.cell(row=row, column=6).value = sup.finaly_row[0].x
            sheet.cell(row=row, column=7).value = sup.finaly_row[0].y
            sheet.cell(row=row, column=8).value = sup.finaly_row[0].error_rate
            sheet.cell(row=row, column=9).value = sup.finaly_row[0].method_determining_point
            sheet.cell(row=row, column=10).value = sup.finaly_row[0].source
            if sup.finaly_row[0].source != "ЕГРН":   
                sheet.cell(row=row, column=11).value = "Новая"
                for number in range(6,11):
                    sheet.cell(row=row, column=number).font = Font(bold=True)
            else:
                sheet.cell(row=row, column=11).value = ""
            if fin.adjacent_codastres:
                    column = 12
                    for adjacent in fin.adjacent_codastres:
                        sheet.cell(row=row, column=column).value = adjacent
                        column+=1
            row+=1
                
    if os.path.isfile(path):
        try:  
            workbook.save(path)
        except PermissionError:
            return False
    else:
        current_datetime = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f"/new_table_{current_datetime}.xlsx"
        workbook.save(path+file_name)
        
    workbook.close()
    
    return True