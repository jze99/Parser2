from Data import GeodesyObject, NewGeodesyObject
import openpyxl


def ReturnList(path:str):
    workbook = openpyxl.load_workbook(path)   
    sheets = workbook.sheetnames
    workbook.close()
    return sheets

def UnploadingGeodezia(path:str, _sheet:str):
    
    NewGeodesyObject.clear()
    
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[_sheet]
    
    # Получаем все строки в листе
    rows = sheet.iter_rows(values_only=True)
    # Пропускаем первую строку
    next(rows)
    
    NewGeodesyObject.clear()
    
    for row in rows:
        if row[0] == None:
            workbook.close()
            return True
        NewGeodesyObject.append(
            GeodesyObject(
                x=float(row[0]),
                y=float(row[1]),
                errorRate=float(row[2]),
                methodDetermination=row[3],
                source=row[4] 
            )
        )
    
    workbook.close()
    return True