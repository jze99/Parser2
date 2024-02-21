import openpyxl
from Data import KPTCadObject, NewKPTCadObject

def UnploadKPT(path:str):
    
    NewKPTCadObject.clear()
    
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['2']
    
    row_presence_coordinates = None
    row_error_determining_area = None
    
    # Проходим по всем ячейкам в первой строке
    for icell, cell in enumerate(sheet[1]):
        if cell.value:
            if cell.value == "Кадастровый номер" or cell.value == "кадастровый номер":
                row_cad_number = icell
                continue
                
            if cell.value == "вид объекта (ОКС, ЗУ)" or cell.value == "Вид ОН (ЗУ, ОКС)" or cell.value == "вид ОН (ОКС, ЗУ)":
                row_type_of_object = icell
                continue
                
            if cell.value == "Площадь" or cell.value == "Площадь КПТ":
                row_square = icell
                continue
                
            if cell.value == "Погрешность определения площади":
                row_error_determining_area = icell
                continue
            if cell.value == "Наличие координат":
                row_presence_coordinates = icell
                continue
                
    # Получаем все строки в листе
    rows = sheet.iter_rows(values_only=True)
    # Пропускаем первую строку
    next(rows)            
                
    for row in rows:
        if row_error_determining_area == None and row_presence_coordinates == None:
            NewKPTCadObject.append(
                KPTCadObject(
                    CadNumber=row[row_cad_number],
                    type_of_object=row[row_type_of_object],
                    square=float(row[row_square]),
                    error_determining_area="данные отсутствуют",
                    presence_coordinates="данные отсутствуют",
                )
            ) 
        if row_error_determining_area != None and row_presence_coordinates == None:
            NewKPTCadObject.append(
                KPTCadObject(
                    CadNumber=row[row_cad_number],
                    type_of_object=row[row_type_of_object],
                    square=float(row[row_square]),
                    error_determining_area=row[row_error_determining_area],
                    presence_coordinates="данные отсутствуют",
                )
            )
        if row_error_determining_area == None and row_presence_coordinates != None:
            NewKPTCadObject.append(
                KPTCadObject(
                    CadNumber=row[row_cad_number],
                    type_of_object=row[row_type_of_object],
                    square=float(row[row_square]),
                    error_determining_area="данные отсутствуют",
                    presence_coordinates=row[row_presence_coordinates],
                )
            )
        if row_error_determining_area != None and row_presence_coordinates != None:
            NewKPTCadObject.append(
                KPTCadObject(
                    CadNumber=row[row_cad_number],
                    type_of_object=row[row_type_of_object],
                    square=float(row[row_square]),
                    error_determining_area=row[row_error_determining_area],
                    presence_coordinates=row[row_presence_coordinates],
                )
            )       
        
    
    sheet = workbook['1'] 
    
    for icell, cell in enumerate(sheet[1]):
        if cell.value:
            if cell.value == "Кадастровый номер":
                row_cad_number = icell
                continue
            if cell.value == "Номер точки":
                row_point_number = icell
                continue
                
            if cell.value == "X":
                row_x = icell
                continue
                
            if cell.value == "У":
                row_y = icell
                continue
                
            if cell.value == "Погрешность" or cell.value == "Погршеность точки":
                row_error_rate = icell
                continue
                
            if cell.value == "Метод определения точки":
                row_method_determining_point = icell
                continue
            if cell.value == "Источник":
                row_source = icell
                continue
                 
    rows = sheet.iter_rows(values_only=True)
    next(rows)
    cad_number_to_row = {}
    
    for row in rows:
        cad_number = row[row_cad_number]
        if cad_number not in cad_number_to_row:
            cad_number_to_row[cad_number] = []
        cad_number_to_row[cad_number].append(row)   
            
    for data in NewKPTCadObject:
        
        if data.presence_coordinates == 'нет координат в ЕГРН':
            continue
        
        if data.CadNumber in cad_number_to_row:
            row_data = cad_number_to_row[data.CadNumber]
            for row in row_data:
                if len(row) > 7:
                    data.AddData(
                        point_number=row[2],
                        x=row[3],
                        y=row[4],
                        error_rate=row[5],
                        method_determining_point=row[6],
                        source=row[7]
                    )
                elif len(row) == 7:
                    data.AddData(
                        x=row[2],
                        y=row[3],
                        error_rate=row[4],
                        method_determining_point=row[5],
                        source= row[6]
                    )
        
        if data.presence_coordinates == "данные отсутствуют":
            if data.x:
                data.presence_coordinates = "есть координаты в ЕГРН"
                if data.point_number[0] == 0:
                    for ipoint in range(len(data.point_number)):
                        data.point_number[ipoint] = ipoint + 1
            else:
                data.presence_coordinates = "нет координат в ЕГРН"
    
    
    # Закрываем файл Excel
    workbook.close()
    
    return True
        